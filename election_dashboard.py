#!/usr/bin/env python3
"""
FCT 2026 Area Council Elections - Live Dashboard
Scrapes INEC IReV API every 3 minutes and serves an interactive dashboard.

Usage:
    python3 election_dashboard.py

Dashboard: http://localhost:5050
"""

import csv
import io
import json
import os
import re
import sqlite3
import threading
import time
import traceback
from datetime import datetime
from queue import Queue

import openpyxl
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from flask import Flask, Response, jsonify, request, send_file
from flask_cors import CORS

# OCR imports
try:
    import pytesseract
    from PIL import Image, ImageFilter, ImageEnhance, ImageOps
    OCR_AVAILABLE = True
    print("[OCR] Tesseract + Pillow loaded successfully")
except ImportError:
    OCR_AVAILABLE = False
    print("[OCR] pytesseract/Pillow not installed – OCR disabled")

# ─── Config ───────────────────────────────────────────────────────────────────
API_BASE = "https://dolphin-app-sleqh.ondigitalocean.app/api/v1"
API_KEY = "4SXkHM7Amb1SbF4C8do6816dmbbwqPp7akRbrmcV"
FCT_STATE_ID = 15
SCRAPE_INTERVAL = 120  # 2 minutes — faster refresh for live election day
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "election_data.db")
EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "FCT_2026_Area_Council_Elections.xlsx")

ELECTION_TYPE_IDS = {
    "CHAIRMAN": "5f129a04df41d910dcdc1d55",
    "COUNCILLOR": "5f129a04df41d910dcdc1d56",
}

app = Flask(__name__)
_frontend_url = os.getenv("FRONTEND_URL", "")
_cors_origins = [o for o in [
    _frontend_url,
    "http://localhost:3000",
    "http://localhost:3001",
    "http://127.0.0.1:3000",
    "http://127.0.0.1:3001",
] if o]
CORS(app, resources={r"/api/*": {"origins": _cors_origins}})

scraper_status = {
    "last_scrape": None, "status": "idle", "error": None,
    "scrape_count": 0, "message": ""
}
# SSE clients queue
sse_clients: list[Queue] = []
sse_clients_lock = threading.Lock()

# ─── HTTP Session with retries ───────────────────────────────────────────────
_session = None

def get_session():
    global _session
    if _session is None:
        _session = requests.Session()
        retries = Retry(
            total=4,
            backoff_factor=5,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=["GET"],
            raise_on_status=False,
        )
        adapter = HTTPAdapter(
            max_retries=retries,
            pool_connections=10,
            pool_maxsize=10,
            pool_block=False,
        )
        _session.mount("https://", adapter)
        _session.mount("http://", adapter)
        _session.headers.update({
            "x-api-key": API_KEY,
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "en-US,en;q=0.9",
            "Origin": "https://www.inecelectionresults.ng",
            "Referer": "https://www.inecelectionresults.ng/",
            "Connection": "keep-alive",
        })
    _session.headers["x-api-rt"] = str(int(time.time() * 1000))
    return _session


# ─── Database ─────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=30000")
    return conn


def init_db():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS elections (
            id TEXT PRIMARY KEY,
            full_name TEXT,
            election_type TEXT,
            election_date TEXT,
            state_name TEXT,
            domain_name TEXT,
            total_pus INTEGER DEFAULT 0,
            total_results INTEGER DEFAULT 0,
            pct REAL DEFAULT 0,
            raw_json TEXT,
            first_seen TEXT DEFAULT (datetime('now')),
            last_updated TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS lgas (
            id TEXT PRIMARY KEY,
            election_id TEXT,
            lga_name TEXT,
            lga_code TEXT,
            lga_id INTEGER,
            state_name TEXT,
            total_wards INTEGER DEFAULT 0,
            total_pus INTEGER DEFAULT 0,
            results_uploaded INTEGER DEFAULT 0,
            raw_json TEXT,
            last_updated TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (election_id) REFERENCES elections(id)
        );

        CREATE TABLE IF NOT EXISTS wards (
            id TEXT PRIMARY KEY,
            lga_db_id TEXT,
            election_id TEXT,
            ward_name TEXT,
            ward_code TEXT,
            ward_id INTEGER,
            lga_name TEXT,
            total_pus INTEGER DEFAULT 0,
            results_uploaded INTEGER DEFAULT 0,
            raw_json TEXT,
            last_updated TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (lga_db_id) REFERENCES lgas(id),
            FOREIGN KEY (election_id) REFERENCES elections(id)
        );

        CREATE TABLE IF NOT EXISTS polling_units (
            id TEXT PRIMARY KEY,
            election_id TEXT,
            ward_id TEXT,
            lga_name TEXT,
            ward_name TEXT,
            pu_name TEXT,
            pu_code TEXT,
            has_result INTEGER DEFAULT 0,
            document_url TEXT,
            document_size INTEGER DEFAULT 0,
            result_uploaded_at TEXT,
            raw_json TEXT,
            last_updated TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (election_id) REFERENCES elections(id)
        );

        CREATE TABLE IF NOT EXISTS scrape_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT DEFAULT (datetime('now')),
            election_type TEXT,
            total_pus INTEGER,
            results_uploaded INTEGER,
            percentage REAL,
            breakdown TEXT
        );

        CREATE TABLE IF NOT EXISTS candidates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            area_council TEXT,
            candidate_name TEXT,
            party_full TEXT,
            party_abbrev TEXT,
            status TEXT,
            gender TEXT,
            notes TEXT,
            position_type TEXT
        );

        CREATE TABLE IF NOT EXISTS area_councils (
            name TEXT PRIMARY KEY,
            total_wards INTEGER,
            polling_units INTEGER,
            registered_voters TEXT,
            chairmanship_candidates INTEGER,
            councillorship_positions INTEGER
        );

        CREATE TABLE IF NOT EXISTS ocr_results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pu_id TEXT UNIQUE,
            pu_code TEXT,
            pu_name TEXT,
            ward_name TEXT,
            lga_name TEXT,
            election_id TEXT,
            registered_voters INTEGER DEFAULT 0,
            accredited_voters INTEGER DEFAULT 0,
            total_valid_votes INTEGER DEFAULT 0,
            total_rejected_votes INTEGER DEFAULT 0,
            party_votes TEXT DEFAULT '{}',
            ocr_confidence REAL DEFAULT 0,
            ocr_raw_text TEXT,
            document_url TEXT,
            processed_at TEXT DEFAULT (datetime('now')),
            status TEXT DEFAULT 'pending',
            FOREIGN KEY (pu_id) REFERENCES polling_units(id),
            FOREIGN KEY (election_id) REFERENCES elections(id)
        );
    """)
    conn.commit()
    conn.close()


# ─── Excel Data Loader ───────────────────────────────────────────────────────
def load_excel_data():
    if not os.path.exists(EXCEL_PATH):
        print(f"[WARN] Excel file not found: {EXCEL_PATH}")
        return

    conn = get_db()
    wb = openpyxl.load_workbook(EXCEL_PATH)

    ws = wb["Election Overview"]
    for row in ws.iter_rows(min_row=5, max_row=10, values_only=True):
        if row[0] and row[0] != "TOTAL":
            conn.execute("""
                INSERT OR REPLACE INTO area_councils
                (name, total_wards, polling_units, registered_voters,
                 chairmanship_candidates, councillorship_positions)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (row[0], row[1], row[2], str(row[3]) if row[3] else "N/A", row[4], row[5]))

    ws = wb["Chairmanship Candidates"]
    conn.execute("DELETE FROM candidates WHERE position_type='Chairmanship'")
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0] and isinstance(row[0], int) and row[2]:
            conn.execute("""
                INSERT INTO candidates
                (area_council, candidate_name, party_full, party_abbrev,
                 status, gender, notes, position_type)
                VALUES (?, ?, ?, ?, ?, ?, ?, 'Chairmanship')
            """, (row[1], row[2], row[3], row[4], row[5], row[6], row[7]))

    conn.commit()
    conn.close()
    print("[OK] Excel data loaded")


# ─── API Calls ────────────────────────────────────────────────────────────────
def api_get(endpoint, params=None, timeout=90, retries=2):
    """Fetch from INEC API with retry logic and connection keep-alive."""
    session = get_session()
    url = f"{API_BASE}/{endpoint}"
    last_err = None

    for attempt in range(retries + 1):
        try:
            resp = session.get(url, params=params, timeout=(15, timeout))
            resp.raise_for_status()
            return resp.json()
        except requests.exceptions.Timeout:
            last_err = "Timeout"
            print(f"[API TIMEOUT] {url} (attempt {attempt+1}/{retries+1})")
            if attempt < retries:
                time.sleep(3 * (attempt + 1))
        except requests.exceptions.ConnectionError as e:
            last_err = str(e)
            print(f"[API CONN ERR] {url}: {e} (attempt {attempt+1}/{retries+1})")
            # Reset session on connection errors to get fresh TCP connection
            if attempt < retries:
                global _session
                _session = None
                time.sleep(5 * (attempt + 1))
        except Exception as e:
            last_err = str(e)
            print(f"[API ERROR] {url}: {e} (attempt {attempt+1}/{retries+1})")
            if attempt < retries:
                time.sleep(2)

    print(f"[API FAILED] {url} after {retries+1} attempts: {last_err}")
    return None


# ─── Scraper: Phase 1 - Quick Stats ──────────────────────────────────────────
def discover_elections():
    """Find all FCT 2026 elections."""
    elections = {"CHAIRMAN": [], "COUNCILLOR": []}

    for etype, type_id in ELECTION_TYPE_IDS.items():
        scraper_status["message"] = f"Discovering {etype} elections..."
        data = api_get("elections", {"election_type": type_id})
        if not data or "data" not in data:
            continue

        for e in (data.get("data") or []):
            state = e.get("state", {})
            state_id = state.get("state_id") if state else None
            edate = e.get("election_date", "")
            # Only FCT + 2026 elections
            if state_id == FCT_STATE_ID and "2026" in edate:
                elections[etype].append({
                    "id": e["_id"],
                    "full_name": e.get("full_name", ""),
                    "election_date": edate,
                    "domain_name": e.get("domain", {}).get("name", ""),
                    "raw": e,
                })
        time.sleep(1)

    return elections


def scrape_stats(elections):
    """Phase 1: Get quick result/stats for each election."""
    conn = get_db()

    for etype, elist in elections.items():
        total_pus_all = 0
        total_results_all = 0
        breakdown = {}

        for e in elist:
            eid = e["id"]
            name = e["domain_name"] or e["full_name"]
            scraper_status["message"] = f"Stats: {name}"

            # Save election
            conn.execute("""
                INSERT OR IGNORE INTO elections
                (id, full_name, election_type, election_date, state_name, domain_name, raw_json)
                VALUES (?, ?, ?, ?, 'FCT', ?, ?)
            """, (eid, e["full_name"], etype, e["election_date"],
                  e["domain_name"], json.dumps(e["raw"])))

            # Get stats
            stats = api_get(f"elections/{eid}/result/stats", timeout=90)
            if stats and "data" in stats:
                sd = stats["data"]
                total_pus = sd.get("expected", sd.get("pus", 0))
                total_docs = sd.get("documents", 0)
                pct = (total_docs / total_pus * 100) if total_pus > 0 else 0

                conn.execute("""
                    UPDATE elections SET total_pus=?, total_results=?, pct=?,
                    last_updated=datetime('now') WHERE id=?
                """, (total_pus, total_docs, round(pct, 1), eid))

                total_pus_all += total_pus
                total_results_all += total_docs
                breakdown[name] = {
                    "pus": total_pus, "results": total_docs,
                    "pct": round(pct, 1)
                }

                print(f"  [{etype}] {name}: {total_docs}/{total_pus} ({pct:.1f}%)")
            time.sleep(0.5)

        # Log
        pct_all = (total_results_all / total_pus_all * 100) if total_pus_all > 0 else 0
        conn.execute("""
            INSERT INTO scrape_log (election_type, total_pus, results_uploaded, percentage, breakdown)
            VALUES (?, ?, ?, ?, ?)
        """, (etype, total_pus_all, total_results_all, round(pct_all, 1), json.dumps(breakdown)))

    conn.commit()
    conn.close()


def scrape_lga_detail(elections):
    """Phase 2: Get LGA/ward detail for ALL elections (chairman + councillor)."""
    conn = get_db()

    for etype in ["CHAIRMAN", "COUNCILLOR"]:
        for e in elections.get(etype, []):
            eid = e["id"]
            name = e["domain_name"] or e["full_name"]
            scraper_status["message"] = f"LGA detail: {name}"

            lga_data = api_get(f"elections/{eid}/lga/state/{FCT_STATE_ID}", timeout=90)
            if not lga_data or "data" not in lga_data:
                time.sleep(1)
                continue

            for lga_entry in (lga_data.get("data") or []):
                lga = lga_entry.get("lga", {})
                lga_db_id = lga_entry.get("_id", "")
                lga_name = lga.get("name", name)
                wards = lga_entry.get("wards", [])

                conn.execute("""
                    INSERT OR REPLACE INTO lgas
                    (id, election_id, lga_name, lga_code, lga_id, state_name,
                     total_wards, raw_json, last_updated)
                    VALUES (?, ?, ?, ?, ?, 'FCT', ?, ?, datetime('now'))
                """, (lga_db_id, eid, lga_name, lga.get("code", ""),
                      lga.get("lga_id", 0), len(wards), json.dumps(lga_entry)))

                for ward in wards:
                    conn.execute("""
                        INSERT OR REPLACE INTO wards
                        (id, lga_db_id, election_id, ward_name, ward_code,
                         ward_id, lga_name, raw_json, last_updated)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
                    """, (ward.get("_id", ""), lga_db_id, eid,
                          ward.get("name", ""), ward.get("code", ""),
                          ward.get("ward_id", 0), lga_name, json.dumps(ward)))

            time.sleep(1)

    conn.commit()
    conn.close()


def scrape_ward_pus(elections):
    """Phase 3: Get polling unit results per ward for ALL elections (chairman + councillor)."""
    scrape_ward_pus_by_type(elections, "CHAIRMAN")
    scrape_ward_pus_by_type(elections, "COUNCILLOR")


def scrape_ward_pus_by_type(elections, etype):
    """Phase 3: Get polling unit results per ward for a specific election type."""
    conn = get_db()

    for e in elections.get(etype, []):
        eid = e["id"]
        name = e["domain_name"] or e["full_name"]
        wards = conn.execute(
            "SELECT id, ward_name, lga_name FROM wards WHERE election_id=?", (eid,)
        ).fetchall()

        lga_pu_counts = {}
        lga_result_counts = {}

        for ward in wards:
            wid = ward["id"]
            wname = ward["ward_name"]
            lname = ward["lga_name"]
            scraper_status["message"] = f"PUs: {lname}/{wname} [{etype[:5]}]"

            pu_data = api_get(f"elections/{eid}/pus", {"ward": wid}, timeout=90)
            if not pu_data or "data" not in pu_data:
                time.sleep(1)
                continue

            ward_pus = 0
            ward_results = 0
            for pu in (pu_data.get("data") or []):
                ward_pus += 1
                doc = pu.get("document", {})
                has_result = 1 if (doc and doc.get("url")) else 0
                if has_result:
                    ward_results += 1

                conn.execute("""
                    INSERT OR REPLACE INTO polling_units
                    (id, election_id, ward_id, lga_name, ward_name, pu_name,
                     pu_code, has_result, document_url, document_size,
                     result_uploaded_at, raw_json, last_updated)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
                """, (pu.get("_id", ""), eid, wid, lname, wname,
                      pu.get("name", ""), pu.get("pu_code", ""),
                      has_result,
                      doc.get("url", "") if doc else "",
                      doc.get("size", 0) if doc else 0,
                      doc.get("updated_at", "") if doc else "",
                      json.dumps(pu)))

            conn.execute("""
                UPDATE wards SET total_pus=?, results_uploaded=?,
                last_updated=datetime('now') WHERE id=? AND election_id=?
            """, (ward_pus, ward_results, wid, eid))

            lga_pu_counts[lname] = lga_pu_counts.get(lname, 0) + ward_pus
            lga_result_counts[lname] = lga_result_counts.get(lname, 0) + ward_results
            time.sleep(0.3)  # Slightly faster for election day

        # Update LGA totals
        for lname in lga_pu_counts:
            conn.execute("""
                UPDATE lgas SET total_pus=?, results_uploaded=?,
                last_updated=datetime('now')
                WHERE lga_name=? AND election_id=?
            """, (lga_pu_counts[lname], lga_result_counts[lname], lname, eid))

    conn.commit()
    conn.close()


# ─── OCR Processing Pipeline ─────────────────────────────────────────────────
# Known INEC result sheet party abbreviations (in order they appear on form)
INEC_PARTIES = [
    "A", "AA", "AAC", "ADC", "ADP", "APC", "APGA", "APM", "APP",
    "BP", "LP", "NNPP", "NRM", "PDP", "PRP", "SDP", "YPP", "ZLP"
]
# All registered parties for INEC 2026 elections (full set for OCR matching)
MAJOR_PARTIES = ["APC", "PDP", "LP", "NNPP", "SDP", "ADC", "APGA", "AA", "ADP", "APM", "ZLP",
                 "YPP", "NRM", "BOOT", "BP", "APP", "PRP", "AAC", "Action Alliance"]

# Common OCR misreads of party names (handwriting variations)
PARTY_ALIASES = {
    "A.P.C": "APC", "A P C": "APC", "APC.": "APC", "APPC": "APC",
    "P.D.P": "PDP", "P D P": "PDP", "PDP.": "PDP", "POP": "PDP",
    "L.P": "LP", "L P": "LP", "LP.": "LP",
    "N.N.P.P": "NNPP", "NNPP.": "NNPP",
    "S.D.P": "SDP", "S D P": "SDP", "SDP.": "SDP",
    "A.D.C": "ADC", "ADC.": "ADC",
    "A.P.G.A": "APGA", "APGA.": "APGA",
    "A.A": "AA", "AA.": "AA",
    "A.D.P": "ADP", "ADP.": "ADP",
    "Y.P.P": "YPP", "YPP.": "YPP",
    "A.P.M": "APM", "APM.": "APM",
    "Z.L.P": "ZLP", "ZLP.": "ZLP",
}

OCR_BATCH_SIZE = 15  # Process max N images per cycle (increased for faster processing)
OCR_LOCK = threading.Lock()


def preprocess_image_standard(img):
    """Standard preprocessing for printed text on INEC forms."""
    if img.mode != "L":
        img = img.convert("L")
    enhancer = ImageEnhance.Contrast(img)
    img = enhancer.enhance(2.0)
    img = img.filter(ImageFilter.SHARPEN)
    w, h = img.size
    if w < 2000:
        scale = 2000 / w
        img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
    img = img.point(lambda x: 0 if x < 140 else 255, "1")
    return img


def preprocess_image_handwriting(img):
    """Enhanced preprocessing for handwritten text on INEC result sheets.
    Uses fast numpy-based adaptive thresholding for pen/pencil on paper."""
    import numpy as np
    from PIL import ImageFilter as IF

    if img.mode != "L":
        img = img.convert("L")

    # Auto-rotate based on EXIF if available
    try:
        img = ImageOps.exif_transpose(img)
    except Exception:
        pass

    # Scale up significantly for handwriting (Tesseract works best at 300+ DPI)
    w, h = img.size
    if w < 2500:
        scale = 2500 / w
        img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)

    # Denoise with median filter to remove noise, preserve edges
    img = img.filter(IF.MedianFilter(size=3))

    # Strong contrast for handwriting (ink vs paper)
    enhancer = ImageEnhance.Contrast(img)
    img = enhancer.enhance(2.5)

    # Brightness adjustment (make paper whiter)
    enhancer = ImageEnhance.Brightness(img)
    img = enhancer.enhance(1.3)

    # Sharpen heavily for handwriting strokes
    img = img.filter(IF.SHARPEN)
    img = img.filter(IF.SHARPEN)

    # Fast adaptive thresholding using numpy uniform filter (Sauvola-like)
    arr = np.array(img, dtype=np.float64)
    # Use a fast box filter via cumsum for local mean
    block = 51
    pad = block // 2

    # Padded cumulative sum for fast local mean
    padded = np.pad(arr, pad, mode='edge')
    # Integral image for fast mean calculation
    integral = np.cumsum(np.cumsum(padded, axis=0), axis=1)
    h_arr, w_arr = arr.shape

    # Compute local mean from integral image
    y1 = np.arange(h_arr)
    y2 = y1 + block
    x1 = np.arange(w_arr)
    x2 = x1 + block
    # Use broadcasting to get sums
    local_sum = (integral[np.ix_(y2, x2)] - integral[np.ix_(y1, x2)]
                 - integral[np.ix_(y2, x1)] + integral[np.ix_(y1, x1)])
    local_mean = local_sum / (block * block)

    # Binarize: pixel < local_mean - offset → black, else white
    offset = 15
    binary = np.where(arr < (local_mean - offset), 0, 255).astype(np.uint8)
    img = Image.fromarray(binary)

    return img


def preprocess_image_aggressive(img):
    """Most aggressive preprocessing: maximum contrast + dilation for faded handwriting."""
    if img.mode != "L":
        img = img.convert("L")
    try:
        img = ImageOps.exif_transpose(img)
    except Exception:
        pass
    w, h = img.size
    if w < 2500:
        scale = 2500 / w
        img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
    # Maximum contrast
    img = ImageOps.autocontrast(img, cutoff=5)
    enhancer = ImageEnhance.Contrast(img)
    img = enhancer.enhance(3.0)
    # Heavy sharpen
    img = img.filter(ImageFilter.SHARPEN)
    img = img.filter(ImageFilter.SHARPEN)
    img = img.filter(ImageFilter.EDGE_ENHANCE_MORE)
    # Fixed threshold (aggressive)
    img = img.point(lambda x: 0 if x < 120 else 255, "1")
    return img


def extract_numbers_from_text(text):
    """Extract all number sequences from OCR text, including common OCR errors.
    Handles handwritten digit misreads and comma/period-separated numbers."""
    # Remove commas in numbers like "1,234" → "1234"
    cleaned = re.sub(r'(\d),(\d)', r'\1\2', text)
    # First try clean numbers
    nums = re.findall(r'\b\d{1,7}\b', cleaned)
    if nums:
        return nums

    # Try to fix common OCR misreads in the context of number-like sequences
    # Only apply substitutions near digit-like characters
    ocr_fixes = [
        ("O", "0"), ("o", "0"), ("Q", "0"),
        ("l", "1"), ("I", "1"), ("i", "1"), ("|", "1"),
        ("Z", "2"), ("z", "2"),
        ("E", "3"),
        ("A", "4"),
        ("S", "5"), ("s", "5"),
        ("G", "6"), ("b", "6"),
        ("T", "7"),
        ("B", "8"),
        ("g", "9"), ("q", "9"),
    ]
    # Look for sequences that look like they should be numbers (mixed digits + letters)
    candidates = re.findall(r'[\dOoQlIi|ZzESsGbTBgqA]{1,7}', text)
    result = []
    for cand in candidates:
        fixed = cand
        for old, new in ocr_fixes:
            fixed = fixed.replace(old, new)
        if re.match(r'^\d{1,7}$', fixed):
            result.append(fixed)
    return result or nums


def parse_inec_result_sheet(ocr_text):
    """
    Parse OCR text from an INEC result sheet to extract vote data.
    INEC EC8A forms have a structured layout:
      - Header: INEC logo, election type, date, state, LGA, ward, PU info
      - Column headers: S/N | Party | Score in Figures | Score in Words
      - Body: Numbered rows with party abbreviation | votes in figures | votes in words
      - Footer: Total valid votes, rejected ballots, total votes cast
      - Signatures section at bottom

    Handles both printed (thermal) and handwritten (pen) result sheets.
    """
    result = {
        "registered_voters": 0,
        "accredited_voters": 0,
        "total_valid_votes": 0,
        "total_rejected_votes": 0,
        "party_votes": {},
        "confidence": 0.0,
    }

    if not ocr_text or not ocr_text.strip():
        return result

    lines = ocr_text.strip().split("\n")
    text_upper = ocr_text.upper()

    # ─── 1. Try to find registered / accredited voters ─────────
    for line in lines:
        lu = line.upper().strip()
        nums = extract_numbers_from_text(line)
        if not nums:
            continue
        if any(kw in lu for kw in ["REGISTER", "REG.", "REGIST", "REG VOT", "REG. VOT"]):
            val = int(nums[0])
            if 10 <= val <= 10000:  # Sanity check: valid PU registration range
                result["registered_voters"] = val
        elif any(kw in lu for kw in ["ACCREDIT", "ACCRED", "ACCR", "ACRED"]):
            val = int(nums[0])
            if 0 <= val <= 10000:
                result["accredited_voters"] = val

    # ─── 2. Extract party votes ────────────────────────────────
    # Strategy: Check each line for party names, handle aliases
    party_found = 0
    for line in lines:
        lu = line.upper().strip()
        nums = extract_numbers_from_text(line)
        if not nums:
            continue

        matched_party = None

        # Check direct party matches first
        for party in MAJOR_PARTIES:
            # Match party abbreviation as word boundary or after S/N number
            if re.search(r'(?:^|\b|\d\s+)' + re.escape(party) + r'(?:\b|\s|$|\.)', lu):
                matched_party = party
                break

        # Check aliases if no direct match
        if not matched_party:
            for alias, canonical in PARTY_ALIASES.items():
                if alias.upper() in lu:
                    matched_party = canonical
                    break

        if matched_party:
            # Take the first clean number as vote count
            vote = int(nums[0])
            if 0 <= vote <= 10000:  # Sanity: no PU has > 10K votes for one party
                result["party_votes"][matched_party] = vote
                party_found += 1

    # ─── 3. Extract totals ─────────────────────────────────────
    for line in lines:
        lu = line.upper().strip()
        nums = extract_numbers_from_text(line)
        if not nums:
            continue
        if any(kw in lu for kw in ["TOTAL VALID", "VALID VOTES", "TOTAL OF VALID",
                                    "VALID VOTE", "TOTAL VOTES CAST"]):
            val = int(nums[-1])
            if 0 <= val <= 15000:
                result["total_valid_votes"] = val
        elif any(kw in lu for kw in ["REJECTED", "REJECT", "REJEC"]):
            val = int(nums[-1])
            if 0 <= val <= 5000:
                result["total_rejected_votes"] = val

    # ─── 4. Infer missing totals ──────────────────────────────
    party_sum = sum(result["party_votes"].values())
    if result["total_valid_votes"] == 0 and party_sum > 0:
        result["total_valid_votes"] = party_sum  # Infer from party sum

    # ─── 5. Calculate confidence ───────────────────────────────
    # Score based on how many fields we could extract
    score = 0
    if result["registered_voters"] > 0:
        score += 10
    if result["accredited_voters"] > 0:
        score += 10
    if party_found >= 1:
        score += 15
    if party_found >= 2:
        score += 15
    if party_found >= 4:
        score += 10
    if party_found >= 8:
        score += 5  # Bonus for extracting many parties
    if result["total_valid_votes"] > 0:
        score += 15
    if result["total_rejected_votes"] >= 0 and any(kw in text_upper for kw in ["REJECTED", "REJECT"]):
        score += 5

    # Cross-validation bonuses
    if party_sum > 0 and result["total_valid_votes"] > 0:
        ratio = party_sum / result["total_valid_votes"]
        if 0.9 <= ratio <= 1.1:
            score += 10  # Good consistency bonus
        elif 0.7 <= ratio <= 1.3:
            score += 5   # Partial consistency

    # Sanity check: accredited <= registered
    if (result["accredited_voters"] > 0 and result["registered_voters"] > 0
            and result["accredited_voters"] <= result["registered_voters"]):
        score += 5

    result["confidence"] = min(score, 100)
    return result


def download_result_image(url):
    """Download result sheet image from INEC IReV."""
    try:
        session = get_session()
        resp = session.get(url, timeout=30, stream=True)
        resp.raise_for_status()

        img_data = io.BytesIO(resp.content)
        img = Image.open(img_data)
        return img
    except Exception as e:
        print(f"[OCR] Failed to download image: {e}")
        return None


def process_ocr_batch():
    """Process a batch of unprocessed result images via OCR.
    Uses short DB transactions to avoid locking issues with the scraper thread."""
    if not OCR_AVAILABLE:
        return 0

    with OCR_LOCK:
        # Phase 1: Quick DB read to get list of unprocessed PUs, then close
        conn = get_db()
        unprocessed = conn.execute("""
            SELECT p.id, p.pu_code, p.pu_name, p.ward_name, p.lga_name,
                   p.election_id, p.document_url
            FROM polling_units p
            LEFT JOIN ocr_results o ON p.id = o.pu_id
            WHERE p.has_result = 1
              AND p.document_url != ''
              AND o.pu_id IS NULL
            ORDER BY p.result_uploaded_at DESC
            LIMIT ?
        """, (OCR_BATCH_SIZE,)).fetchall()
        pu_list = [dict(row) for row in unprocessed]
        conn.close()  # Release DB immediately

        processed = 0
        for pu in pu_list:
            doc_url = pu["document_url"]
            if not doc_url:
                continue

            scraper_status["message"] = f"OCR: {pu['lga_name']}/{pu['ward_name']}/{pu['pu_code']}"
            print(f"[OCR] Processing: {pu['pu_code']} ({pu['lga_name']}/{pu['ward_name']})")

            ocr_result = None
            try:
                # Download image (network I/O — no DB lock held)
                img = download_result_image(doc_url)
                if img is None:
                    ocr_result = {
                        "status": "failed", "ocr_raw_text": "Download failed",
                        "ocr_confidence": 0, "parsed": None
                    }
                else:
                    # Multi-pass OCR: try different preprocessing for handwriting
                    best_text = ""
                    best_confidence = 0
                    best_parsed = None

                    passes = [
                        ("standard",    preprocess_image_standard,    "--psm 6 --oem 3"),
                        ("handwriting", preprocess_image_handwriting, "--psm 4 --oem 3"),
                        ("aggressive",  preprocess_image_aggressive,  "--psm 11 --oem 3"),
                        ("raw-scaled",  None,                         "--psm 6 --oem 3 -c tessedit_char_whitelist=0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz /.()-"),
                    ]

                    for pass_name, preprocess_fn, tess_config in passes:
                        try:
                            if preprocess_fn:
                                pimg = preprocess_fn(img.copy())
                            else:
                                # Raw scaled pass
                                pimg = img.copy()
                                if pimg.mode != "L":
                                    pimg = pimg.convert("L")
                                w4, h4 = pimg.size
                                if w4 < 2000:
                                    sc = 2000 / w4
                                    pimg = pimg.resize((int(w4 * sc), int(h4 * sc)), Image.LANCZOS)

                            text = pytesseract.image_to_string(pimg, config=tess_config)
                            parsed = parse_inec_result_sheet(text)
                            if parsed["confidence"] > best_confidence:
                                best_text = text
                                best_confidence = parsed["confidence"]
                                best_parsed = parsed
                        except Exception as pe:
                            print(f"  [WARN] Pass '{pass_name}' failed: {pe}")

                    ocr_text = best_text
                    parsed = best_parsed or parse_inec_result_sheet(ocr_text)

                    ocr_result = {
                        "status": "success" if parsed["confidence"] >= 30 else "low_confidence",
                        "ocr_raw_text": ocr_text[:5000],
                        "ocr_confidence": parsed["confidence"],
                        "parsed": parsed
                    }
                    processed += 1
                    print(f"  [OK] Confidence: {parsed['confidence']}% | Parties: {len(parsed['party_votes'])} | Valid: {parsed['total_valid_votes']}")

            except Exception as e:
                print(f"  [ERR] OCR failed for {pu['pu_code']}: {e}")
                ocr_result = {
                    "status": "error", "ocr_raw_text": str(e)[:500],
                    "ocr_confidence": 0, "parsed": None
                }

            # Phase 2: Quick DB write for this one PU (short transaction)
            try:
                conn2 = get_db()
                if ocr_result["parsed"]:
                    p = ocr_result["parsed"]
                    conn2.execute("""
                        INSERT OR REPLACE INTO ocr_results
                        (pu_id, pu_code, pu_name, ward_name, lga_name, election_id,
                         registered_voters, accredited_voters, total_valid_votes,
                         total_rejected_votes, party_votes, ocr_confidence,
                         ocr_raw_text, document_url, status)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        pu["id"], pu["pu_code"], pu["pu_name"],
                        pu["ward_name"], pu["lga_name"], pu["election_id"],
                        p["registered_voters"], p["accredited_voters"],
                        p["total_valid_votes"], p["total_rejected_votes"],
                        json.dumps(p["party_votes"]), p["confidence"],
                        ocr_result["ocr_raw_text"], doc_url, ocr_result["status"]
                    ))
                else:
                    conn2.execute("""
                        INSERT OR REPLACE INTO ocr_results
                        (pu_id, pu_code, pu_name, ward_name, lga_name, election_id,
                         document_url, status, ocr_raw_text, ocr_confidence)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 0)
                    """, (pu["id"], pu["pu_code"], pu["pu_name"],
                          pu["ward_name"], pu["lga_name"], pu["election_id"],
                          doc_url, ocr_result["status"], ocr_result["ocr_raw_text"]))
                conn2.commit()
                conn2.close()
            except Exception as db_err:
                print(f"  [DB ERR] Could not save OCR result: {db_err}")

            time.sleep(0.5)  # Rate limit between downloads

        return processed


# ─── Main Scraper Loop ───────────────────────────────────────────────────────
def scrape_loop():
    """Background scraper with optimized frequency for election day.
    - Quick stats: every cycle (2 min)
    - LGA/Ward structure: every 2nd cycle (4 min)
    - Chairman PU detail: every 2nd cycle (4 min)
    - Councillor PU detail: every 3rd cycle (6 min)
    - OCR: every 4th cycle (8 min) — low priority since structured data is primary
    """
    cycle = 0
    cached_elections = None

    while True:
        try:
            scraper_status["status"] = "scraping"
            scraper_status["error"] = None
            now = datetime.now().strftime("%H:%M:%S")
            print(f"\n{'='*50}")
            print(f"[SCRAPE] Cycle {cycle+1} at {now}")
            print(f"{'='*50}")

            # Discover elections (cache for 10 cycles)
            if cached_elections is None or cycle % 10 == 0:
                cached_elections = discover_elections()
                chair_count = len(cached_elections.get("CHAIRMAN", []))
                council_count = len(cached_elections.get("COUNCILLOR", []))
                print(f"[FOUND] {chair_count} Chairmanship + {council_count} Councillorship elections")

            # Phase 1: Quick stats (every cycle)
            print("[PHASE 1] Quick stats...")
            scrape_stats(cached_elections)

            # Phase 2: LGA/Ward structure (every 2nd cycle — faster discovery)
            if cycle % 2 == 0:
                print("[PHASE 2] LGA/Ward structure (chairman + councillor)...")
                scrape_lga_detail(cached_elections)

            # Phase 3: Deep PU detail — chairman (every 2nd cycle for freshness)
            if cycle % 2 == 0:
                print("[PHASE 3a] Chairman polling unit details...")
                scraper_status["message"] = "Deep scraping chairman PU results..."
                scrape_ward_pus_by_type(cached_elections, "CHAIRMAN")

            # Phase 3b: Deep PU detail — councillor (every 3rd cycle)
            if cycle % 3 == 0:
                print("[PHASE 3b] Councillor polling unit details...")
                scraper_status["message"] = "Deep scraping councillor PU results..."
                scrape_ward_pus_by_type(cached_elections, "COUNCILLOR")

            # Phase 4: OCR result sheet processing (every 4th cycle — low priority)
            if OCR_AVAILABLE and cycle % 4 == 0:
                print("[PHASE 4] OCR processing result sheets...")
                scraper_status["message"] = "Running OCR on result sheets..."
                ocr_count = process_ocr_batch()
                print(f"[PHASE 4] Processed {ocr_count} result sheets via OCR")
                if ocr_count > 0:
                    sse_broadcast({"event": "ocr_complete", "processed": ocr_count})

            scraper_status["last_scrape"] = datetime.now().isoformat()
            scraper_status["status"] = "idle"
            scraper_status["scrape_count"] += 1
            scraper_status["message"] = "Waiting for next cycle"
            sse_broadcast({"event": "scrape_complete", "scrape_count": scraper_status["scrape_count"]})
            cycle += 1
            print(f"[DONE] Next scrape in {SCRAPE_INTERVAL}s")

        except Exception as e:
            scraper_status["status"] = "error"
            scraper_status["error"] = str(e)
            scraper_status["message"] = f"Error: {e}"
            sse_broadcast({"event": "scrape_error", "error": str(e)})
            print(f"[ERROR] {e}")
            traceback.print_exc()

        time.sleep(SCRAPE_INTERVAL)


# ─── Flask API ────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return send_file(
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "dashboard.html"),
        mimetype="text/html"
    )


@app.route("/api/status")
def api_status():
    return jsonify(scraper_status)


@app.route("/api/overview")
def api_overview():
    conn = get_db()
    councils = [dict(r) for r in conn.execute("SELECT * FROM area_councils").fetchall()]
    elections = [dict(r) for r in conn.execute(
        "SELECT id, full_name, election_type, election_date, domain_name, total_pus, total_results, pct FROM elections ORDER BY election_type, domain_name"
    ).fetchall()]

    # Aggregate stats per election type
    stats = {}
    for etype in ["CHAIRMAN", "COUNCILLOR"]:
        row = conn.execute("""
            SELECT COALESCE(SUM(total_pus),0) as total_pus,
                   COALESCE(SUM(total_results),0) as total_results
            FROM elections WHERE election_type=?
        """, (etype,)).fetchone()
        total = row["total_pus"]
        uploaded = row["total_results"]
        stats[etype] = {
            "total_pus": total,
            "results_uploaded": uploaded,
            "percentage": round(uploaded / total * 100, 1) if total > 0 else 0,
        }

    conn.close()
    return jsonify({
        "area_councils": councils,
        "elections": elections,
        "stats": stats,
        "scraper": scraper_status,
    })


@app.route("/api/elections-detail")
def api_elections_detail():
    """Per-election breakdown (chairmanship per LGA)."""
    conn = get_db()
    rows = conn.execute("""
        SELECT id, full_name, election_type, domain_name,
               total_pus, total_results, pct
        FROM elections
        WHERE election_type='CHAIRMAN'
        ORDER BY domain_name
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/lga-breakdown")
def api_lga_breakdown():
    conn = get_db()
    # Use elections table directly (each chairman election = 1 LGA)
    rows = conn.execute("""
        SELECT domain_name as lga_name, total_pus, total_results, pct
        FROM elections WHERE election_type='CHAIRMAN'
        ORDER BY domain_name
    """).fetchall()

    result = []
    for r in rows:
        r = dict(r)
        result.append({
            "lga_name": r["lga_name"],
            "elections": {
                "CHAIRMAN": {
                    "total_pus": r["total_pus"],
                    "results_uploaded": r["total_results"],
                    "percentage": r["pct"] or 0,
                }
            }
        })

    # Also try to merge councillorship data
    council_rows = conn.execute("""
        SELECT domain_name, total_pus, total_results, pct
        FROM elections WHERE election_type='COUNCILLOR'
    """).fetchall()

    # The councillorship domain_name is the ward name, so aggregate differently
    # We'll show them separately in the dashboard
    council_total_pus = sum(r["total_pus"] or 0 for r in council_rows)
    council_total_res = sum(r["total_results"] or 0 for r in council_rows)

    conn.close()
    return jsonify({
        "lga_data": result,
        "councillorship_summary": {
            "total_wards": len(council_rows),
            "total_pus": council_total_pus,
            "total_results": council_total_res,
            "percentage": round(council_total_res / council_total_pus * 100, 1) if council_total_pus > 0 else 0,
        }
    })


@app.route("/api/ward-breakdown/<lga_name>")
def api_ward_breakdown(lga_name):
    conn = get_db()
    rows = conn.execute("""
        SELECT w.ward_name, w.total_pus, w.results_uploaded,
               w.lga_name, e.election_type
        FROM wards w JOIN elections e ON w.election_id = e.id
        WHERE w.lga_name = ?
        ORDER BY w.ward_name
    """, (lga_name,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/councillorship")
def api_councillorship():
    """All councillorship elections with their stats."""
    conn = get_db()
    rows = conn.execute("""
        SELECT id, domain_name as ward_name, total_pus, total_results, pct
        FROM elections WHERE election_type='COUNCILLOR'
        ORDER BY domain_name
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/candidates")
def api_candidates():
    conn = get_db()
    rows = conn.execute("SELECT * FROM candidates ORDER BY area_council, party_abbrev").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/candidates/<area_council>")
def api_candidates_by_council(area_council):
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM candidates WHERE area_council=? ORDER BY party_abbrev",
        (area_council,)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/timeline")
def api_timeline():
    conn = get_db()
    rows = conn.execute("""
        SELECT timestamp, election_type, total_pus, results_uploaded, percentage, breakdown
        FROM scrape_log ORDER BY timestamp DESC LIMIT 300
    """).fetchall()
    conn.close()
    result = []
    for r in rows:
        r = dict(r)
        r["breakdown"] = json.loads(r["breakdown"]) if r["breakdown"] else {}
        result.append(r)
    return jsonify(result)


@app.route("/api/recent-results")
def api_recent_results():
    conn = get_db()
    rows = conn.execute("""
        SELECT pu_name, pu_code, lga_name, ward_name, has_result,
               document_url, result_uploaded_at
        FROM polling_units WHERE has_result=1
        ORDER BY result_uploaded_at DESC LIMIT 50
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/party-analysis")
def api_party_analysis():
    conn = get_db()
    by_party = conn.execute("""
        SELECT party_abbrev, party_full, COUNT(*) as count,
               SUM(CASE WHEN gender='F' THEN 1 ELSE 0 END) as female,
               SUM(CASE WHEN gender='M' THEN 1 ELSE 0 END) as male,
               SUM(CASE WHEN status LIKE '%WITHDRAWN%' THEN 1 ELSE 0 END) as withdrawn
        FROM candidates GROUP BY party_abbrev ORDER BY count DESC
    """).fetchall()

    by_council = conn.execute("""
        SELECT area_council, COUNT(*) as total,
               COUNT(DISTINCT party_abbrev) as parties,
               SUM(CASE WHEN gender='F' THEN 1 ELSE 0 END) as female
        FROM candidates GROUP BY area_council ORDER BY area_council
    """).fetchall()
    conn.close()
    return jsonify({
        "by_party": [dict(r) for r in by_party],
        "by_council": [dict(r) for r in by_council],
    })


@app.route("/api/analytics/party-race")
def api_party_race():
    """
    Enhanced party-level analytics: race standings, candidate leaderboard,
    competitive analysis, and projections per area council.
    """
    conn = get_db()

    # ─── 1. Candidates by party, filtered to active only ──────────
    candidates = conn.execute("""
        SELECT id, area_council, candidate_name, party_full, party_abbrev,
               status, gender, position_type
        FROM candidates ORDER BY area_council, party_abbrev
    """).fetchall()
    candidates = [dict(c) for c in candidates]

    active_candidates = [c for c in candidates if "WITHDRAWN" not in (c.get("status") or "").upper()]
    withdrawn = [c for c in candidates if "WITHDRAWN" in (c.get("status") or "").upper()]

    # ─── 2. Chairmanship Race — per area council ──────────────────
    chairmanship = [c for c in active_candidates if c.get("position_type") == "Chairmanship"]
    councillorship = [c for c in active_candidates if c.get("position_type") == "Councillorship"]

    # Group chairmanship by area council
    chair_by_council = {}
    for c in chairmanship:
        ac = c["area_council"]
        if ac not in chair_by_council:
            chair_by_council[ac] = []
        chair_by_council[ac].append({
            "candidate_name": c["candidate_name"],
            "party": c["party_abbrev"],
            "party_full": c["party_full"],
            "gender": c["gender"],
        })

    # ─── 3. Party strength index (candidate spread across councils) ─
    party_stats = {}
    for c in active_candidates:
        p = c["party_abbrev"]
        if p not in party_stats:
            party_stats[p] = {
                "party": p,
                "party_full": c["party_full"],
                "total_candidates": 0,
                "chairmanship": 0,
                "councillorship": 0,
                "councils_present": set(),
                "female": 0,
                "male": 0,
            }
        party_stats[p]["total_candidates"] += 1
        if c.get("position_type") == "Chairmanship":
            party_stats[p]["chairmanship"] += 1
        else:
            party_stats[p]["councillorship"] += 1
        party_stats[p]["councils_present"].add(c["area_council"])
        if c.get("gender") == "F":
            party_stats[p]["female"] += 1
        else:
            party_stats[p]["male"] += 1

    # Convert sets to counts for JSON serialization
    party_standings = []
    for p, stats in party_stats.items():
        councils_count = len(stats["councils_present"])
        # Strength index: weighted score based on coverage and candidate count
        strength = (
            stats["chairmanship"] * 3 +      # Chairman seats worth more
            stats["councillorship"] * 1 +     # Councillor seats
            councils_count * 2                # Coverage bonus
        )
        party_standings.append({
            "party": stats["party"],
            "party_full": stats["party_full"],
            "total_candidates": stats["total_candidates"],
            "chairmanship": stats["chairmanship"],
            "councillorship": stats["councillorship"],
            "councils_present": councils_count,
            "councils_total": 6,
            "coverage_pct": round(councils_count / 6 * 100, 1),
            "female": stats["female"],
            "male": stats["male"],
            "strength_index": strength,
        })
    party_standings.sort(key=lambda x: x["strength_index"], reverse=True)

    # ─── 4. Competitiveness per area council ──────────────────────
    council_competitiveness = []
    for ac, cands in chair_by_council.items():
        parties = [c["party"] for c in cands]
        major_parties = [p for p in parties if p in ("APC", "PDP", "LP", "NNPP", "SDP", "ADC")]
        council_competitiveness.append({
            "area_council": ac,
            "total_candidates": len(cands),
            "candidates": cands,
            "major_parties": len(major_parties),
            "is_competitive": len(major_parties) >= 3,
            "parties": parties,
        })
    council_competitiveness.sort(key=lambda x: x["total_candidates"], reverse=True)

    # ─── 5. Councillorship spread ─────────────────────────────────
    councillor_by_party = {}
    for c in councillorship:
        p = c["party_abbrev"]
        if p not in councillor_by_party:
            councillor_by_party[p] = 0
        councillor_by_party[p] += 1
    councillor_spread = [{"party": k, "count": v} for k, v in councillor_by_party.items()]
    councillor_spread.sort(key=lambda x: x["count"], reverse=True)

    # ─── 6. Head-to-head matchups (top parties per council) ───────
    matchups = []
    for ac, cands in chair_by_council.items():
        # Highlight top 3 party battles
        top_parties = [c for c in cands if c["party"] in ("APC", "PDP", "LP", "NNPP", "SDP", "ADC")]
        if len(top_parties) >= 2:
            matchups.append({
                "area_council": ac,
                "contenders": top_parties[:4],  # top 4 for display
                "total_in_race": len(cands),
            })

    # ─── 7. Gender balance scorecard ──────────────────────────────
    total_female = sum(1 for c in active_candidates if c.get("gender") == "F")
    total_male = sum(1 for c in active_candidates if c.get("gender") == "M")
    total_active = len(active_candidates)
    gender_by_position = {
        "chairmanship": {
            "female": sum(1 for c in chairmanship if c.get("gender") == "F"),
            "male": sum(1 for c in chairmanship if c.get("gender") == "M"),
            "total": len(chairmanship),
        },
        "councillorship": {
            "female": sum(1 for c in councillorship if c.get("gender") == "F"),
            "male": sum(1 for c in councillorship if c.get("gender") == "M"),
            "total": len(councillorship),
        },
    }

    # ─── 8. Upload progress per LGA (from elections table) ────────
    elections = conn.execute("""
        SELECT election_type, domain_name, total_pus, total_results, pct
        FROM elections
    """).fetchall()
    election_data = [dict(e) for e in elections]

    conn.close()

    return jsonify({
        "party_standings": party_standings,
        "chairmanship_races": council_competitiveness,
        "councillor_spread": councillor_spread,
        "head_to_head": matchups,
        "gender_scorecard": {
            "total_active": total_active,
            "female": total_female,
            "male": total_male,
            "female_pct": round(total_female / total_active * 100, 1) if total_active > 0 else 0,
            "by_position": gender_by_position,
        },
        "withdrawn_count": len(withdrawn),
        "total_candidates": len(candidates),
        "active_candidates": len(active_candidates),
        "election_progress": election_data,
    })


@app.route("/api/live-results")
def api_live_results():
    """Live vote results extracted from INEC IReV API structured data.
    Uses the vote counts directly from the API (no OCR needed).
    Supports ?type=CHAIRMAN|COUNCILLOR filter (default: all)."""
    conn = get_db()

    # Optional type filter
    etype_filter = request.args.get("type", "").upper()

    if etype_filter in ("CHAIRMAN", "COUNCILLOR"):
        rows = conn.execute("""
            SELECT p.id, p.pu_code, p.pu_name, p.ward_name, p.lga_name,
                   p.election_id, p.raw_json, e.election_type, e.domain_name
            FROM polling_units p
            JOIN elections e ON p.election_id = e.id
            WHERE p.has_result = 1 AND p.raw_json IS NOT NULL
            AND e.election_type = ?
            ORDER BY p.lga_name, p.ward_name, p.pu_code
        """, (etype_filter,)).fetchall()
    else:
        rows = conn.execute("""
            SELECT p.id, p.pu_code, p.pu_name, p.ward_name, p.lga_name,
                   p.election_id, p.raw_json, e.election_type, e.domain_name
            FROM polling_units p
            JOIN elections e ON p.election_id = e.id
            WHERE p.has_result = 1 AND p.raw_json IS NOT NULL
            ORDER BY p.lga_name, p.ward_name, p.pu_code
        """).fetchall()

    results = []
    party_totals = {}
    total_registered = 0
    total_accredited = 0
    total_valid = 0
    total_rejected = 0
    pus_with_votes = 0

    # Track by election type
    by_election = {}  # election_id -> {party_totals, ...}

    for row in rows:
        row = dict(row)
        try:
            rj = json.loads(row["raw_json"]) if row["raw_json"] else {}
        except (json.JSONDecodeError, TypeError):
            continue

        registered = int(rj.get("total_registered", 0) or 0)
        accredited = int(rj.get("total_accredited", 0) or 0)
        valid = int(rj.get("valid_votes", 0) or 0)
        invalid = int(rj.get("invalid_votes", 0) or 0)

        # Parse party votes from API response
        votes_raw = rj.get("votes", "[]")
        party_votes = {}
        if isinstance(votes_raw, str):
            try:
                vote_list = json.loads(votes_raw)
            except:
                vote_list = []
        elif isinstance(votes_raw, list):
            vote_list = votes_raw
        else:
            vote_list = []

        for v in vote_list:
            if isinstance(v, dict):
                party_code = (v.get("party_code", "") or "").upper()
                vote_count = int(v.get("vote", 0) or 0)
                if party_code and vote_count > 0:
                    party_votes[party_code] = vote_count

        if party_votes or valid > 0:
            pus_with_votes += 1

        total_registered += registered
        total_accredited += accredited
        total_valid += valid
        total_rejected += invalid

        # Aggregate party totals
        for party, votes in party_votes.items():
            party_totals[party] = party_totals.get(party, 0) + votes

        results.append({
            "pu_code": row["pu_code"],
            "pu_name": row["pu_name"],
            "ward_name": row["ward_name"],
            "lga_name": row["lga_name"],
            "election_type": row.get("election_type", "CHAIRMAN"),
            "domain_name": row.get("domain_name", ""),
            "registered_voters": registered,
            "accredited_voters": accredited,
            "total_valid_votes": valid,
            "total_rejected_votes": invalid,
            "party_votes": party_votes,
        })

        # Track by election
        eid = row.get("election_id", "unknown")
        if eid not in by_election:
            by_election[eid] = {"party_totals": {}, "pus": 0, "valid": 0,
                                "type": row.get("election_type", "CHAIRMAN"),
                                "domain": row.get("domain_name", "")}
        by_election[eid]["pus"] += 1
        by_election[eid]["valid"] += valid
        for party, votes in party_votes.items():
            by_election[eid]["party_totals"][party] = by_election[eid]["party_totals"].get(party, 0) + votes

    # Sort parties by total votes descending
    party_standings = [
        {"party": p, "votes": v}
        for p, v in sorted(party_totals.items(), key=lambda x: x[1], reverse=True)
        if v > 0
    ]

    # Ward-level aggregation
    ward_agg = {}
    for r in results:
        wkey = f"{r['lga_name']}|{r['ward_name']}"
        if wkey not in ward_agg:
            ward_agg[wkey] = {
                "ward_name": r["ward_name"],
                "lga_name": r["lga_name"],
                "pu_count": 0,
                "total_valid": 0,
                "total_registered": 0,
                "total_accredited": 0,
                "party_votes": {},
            }
        ward_agg[wkey]["pu_count"] += 1
        ward_agg[wkey]["total_valid"] += r["total_valid_votes"]
        ward_agg[wkey]["total_registered"] += r["registered_voters"]
        ward_agg[wkey]["total_accredited"] += r["accredited_voters"]
        for party, votes in r["party_votes"].items():
            ward_agg[wkey]["party_votes"][party] = ward_agg[wkey]["party_votes"].get(party, 0) + votes

    # Determine leading party per ward
    ward_results = []
    for w in ward_agg.values():
        leading = max(w["party_votes"].items(), key=lambda x: x[1]) if w["party_votes"] else ("N/A", 0)
        turnout = round(w["total_accredited"] / w["total_registered"] * 100, 1) if w["total_registered"] > 0 else 0
        ward_results.append({
            **w,
            "leading_party": leading[0],
            "leading_votes": leading[1],
            "turnout_pct": turnout,
        })
    ward_results.sort(key=lambda x: (x["lga_name"], x["ward_name"]))

    # LGA-level aggregation
    lga_agg = {}
    for r in results:
        lga = r["lga_name"]
        if lga not in lga_agg:
            lga_agg[lga] = {
                "lga_name": lga, "pu_count": 0, "total_valid": 0,
                "total_registered": 0, "total_accredited": 0,
                "party_votes": {},
            }
        lga_agg[lga]["pu_count"] += 1
        lga_agg[lga]["total_valid"] += r["total_valid_votes"]
        lga_agg[lga]["total_registered"] += r["registered_voters"]
        lga_agg[lga]["total_accredited"] += r["accredited_voters"]
        for party, votes in r["party_votes"].items():
            lga_agg[lga]["party_votes"][party] = lga_agg[lga]["party_votes"].get(party, 0) + votes

    lga_results = []
    for la in lga_agg.values():
        leading = max(la["party_votes"].items(), key=lambda x: x[1]) if la["party_votes"] else ("N/A", 0)
        turnout = round(la["total_accredited"] / la["total_registered"] * 100, 1) if la["total_registered"] > 0 else 0
        top3 = sorted(la["party_votes"].items(), key=lambda x: -x[1])[:3]
        lga_results.append({
            **la,
            "leading_party": leading[0],
            "leading_votes": leading[1],
            "turnout_pct": turnout,
            "top3": [{"party": p, "votes": v} for p, v in top3],
        })
    lga_results.sort(key=lambda x: x["lga_name"])

    total_pus = conn.execute("SELECT COUNT(*) as c FROM polling_units").fetchone()["c"]
    total_with_results = conn.execute("SELECT COUNT(*) as c FROM polling_units WHERE has_result=1").fetchone()["c"]

    conn.close()

    return jsonify({
        "pu_results": results[:200],  # Limit PU-level detail to avoid huge response
        "party_standings": party_standings,
        "ward_results": ward_results,
        "lga_results": lga_results,
        "summary": {
            "total_registered": total_registered,
            "total_accredited": total_accredited,
            "total_valid": total_valid,
            "total_rejected": total_rejected,
            "pus_with_votes": pus_with_votes,
            "pus_with_results": total_with_results,
            "total_pus": total_pus,
            "coverage_pct": round(total_with_results / total_pus * 100, 1) if total_pus > 0 else 0,
            "data_source": "INEC IReV API (structured data)",
        },
    })


# Mapping from INEC IReV LGA names to our area council names
LGA_TO_COUNCIL = {
    "MUNICIPAL": "AMAC",
    "ABAJI": "Abaji",
    "BWARI": "Bwari",
    "GWAGWALADA": "Gwagwalada",
    "KUJE": "Kuje",
    "KWALI": "Kwali",
}

COUNCIL_TO_LGA = {v.upper(): k for k, v in LGA_TO_COUNCIL.items()}


@app.route("/api/chairmanship-race")
def api_chairmanship_race():
    """Live chairmanship race: maps vote data to actual candidates per area council.
    Shows who is winning each Area Council chairmanship position."""
    conn = get_db()

    # Get all chairmanship candidates
    candidates = conn.execute("""
        SELECT candidate_name, party_abbrev, party_full, area_council, status, gender, notes
        FROM candidates
        WHERE position_type = 'Chairmanship'
        ORDER BY area_council, party_abbrev
    """).fetchall()
    candidates = [dict(c) for c in candidates]

    # Get vote data aggregated by LGA and party
    rows = conn.execute("""
        SELECT lga_name, raw_json
        FROM polling_units
        WHERE has_result = 1 AND raw_json IS NOT NULL
    """).fetchall()

    # Aggregate votes by LGA → party
    lga_votes = {}
    lga_stats = {}
    for row in rows:
        row = dict(row)
        lga = row["lga_name"]
        try:
            rj = json.loads(row["raw_json"]) if row["raw_json"] else {}
        except:
            continue

        if lga not in lga_votes:
            lga_votes[lga] = {}
            lga_stats[lga] = {"pus": 0, "registered": 0, "accredited": 0, "valid": 0}

        lga_stats[lga]["pus"] += 1
        lga_stats[lga]["registered"] += int(rj.get("total_registered", 0) or 0)
        lga_stats[lga]["accredited"] += int(rj.get("total_accredited", 0) or 0)
        lga_stats[lga]["valid"] += int(rj.get("valid_votes", 0) or 0)

        votes_raw = rj.get("votes", "[]")
        if isinstance(votes_raw, str):
            try:
                vote_list = json.loads(votes_raw)
            except:
                vote_list = []
        elif isinstance(votes_raw, list):
            vote_list = votes_raw
        else:
            vote_list = []

        for v in vote_list:
            if isinstance(v, dict):
                party_code = (v.get("party_code", "") or "").upper()
                vote_count = int(v.get("vote", 0) or 0)
                if party_code and vote_count > 0:
                    lga_votes[lga][party_code] = lga_votes[lga].get(party_code, 0) + vote_count

    # Build race results per area council
    races = []
    for council_name in ["AMAC", "Abaji", "Bwari", "Gwagwalada", "Kuje", "Kwali"]:
        lga_key = COUNCIL_TO_LGA.get(council_name.upper(), council_name.upper())
        council_candidates = [c for c in candidates if c["area_council"].upper() == council_name.upper()]
        vote_data = lga_votes.get(lga_key, {})
        stats = lga_stats.get(lga_key, {"pus": 0, "registered": 0, "accredited": 0, "valid": 0})

        # Map votes to candidates
        candidate_results = []
        for cand in council_candidates:
            party = cand["party_abbrev"].upper()
            votes = vote_data.get(party, 0)
            vote_pct = round(votes / stats["valid"] * 100, 1) if stats["valid"] > 0 else 0
            candidate_results.append({
                "candidate_name": cand["candidate_name"],
                "party": party,
                "party_full": cand["party_full"],
                "votes": votes,
                "vote_pct": vote_pct,
                "status": cand["status"],
                "gender": cand["gender"],
                "notes": cand.get("notes", ""),
            })

        # Sort by votes descending
        candidate_results.sort(key=lambda x: -x["votes"])

        # Determine lead margin
        if len(candidate_results) >= 2 and candidate_results[0]["votes"] > 0:
            margin = candidate_results[0]["votes"] - candidate_results[1]["votes"]
        else:
            margin = 0

        turnout = round(stats["accredited"] / stats["registered"] * 100, 1) if stats["registered"] > 0 else 0

        races.append({
            "area_council": council_name,
            "candidates": candidate_results,
            "total_pus_counted": stats["pus"],
            "total_valid_votes": stats["valid"],
            "total_registered": stats["registered"],
            "total_accredited": stats["accredited"],
            "turnout_pct": turnout,
            "margin": margin,
            "winner": candidate_results[0] if candidate_results and candidate_results[0]["votes"] > 0 else None,
        })

    conn.close()

    return jsonify({
        "races": races,
        "total_councils": len(races),
        "councils_with_data": len([r for r in races if r["total_valid_votes"] > 0]),
    })


@app.route("/api/councillorship-race")
def api_councillorship_race():
    """Live councillorship race: ward-level results showing which parties lead each ward.
    Each councillor election is per-ward (1 ward = 1 councillorship seat)."""
    conn = get_db()

    # Get all councillor elections with their domain_name (ward name)
    councillor_elections = conn.execute("""
        SELECT id, full_name, domain_name, total_pus, total_results, pct
        FROM elections
        WHERE election_type = 'COUNCILLOR'
        ORDER BY domain_name
    """).fetchall()
    councillor_elections = [dict(e) for e in councillor_elections]

    # Build a map of election_id → ward info
    election_map = {}
    for e in councillor_elections:
        election_map[e["id"]] = e

    # Get all councillor PUs with vote data
    rows = conn.execute("""
        SELECT p.pu_code, p.pu_name, p.ward_name, p.lga_name,
               p.election_id, p.raw_json, e.domain_name
        FROM polling_units p
        JOIN elections e ON p.election_id = e.id
        WHERE e.election_type = 'COUNCILLOR'
        AND p.has_result = 1 AND p.raw_json IS NOT NULL
    """).fetchall()

    # Aggregate votes per ward (each councillor election = 1 ward)
    ward_races = {}  # keyed by election_id (one per ward)
    for row in rows:
        row = dict(row)
        eid = row["election_id"]
        try:
            rj = json.loads(row["raw_json"]) if row["raw_json"] else {}
        except:
            continue

        if eid not in ward_races:
            einfo = election_map.get(eid, {})
            ward_races[eid] = {
                "ward_name": row.get("domain_name") or row["ward_name"],
                "lga_name": row["lga_name"],
                "total_pus_in_ward": einfo.get("total_pus", 0),
                "pus_counted": 0,
                "registered": 0,
                "accredited": 0,
                "valid": 0,
                "rejected": 0,
                "party_votes": {},
            }

        wr = ward_races[eid]
        wr["pus_counted"] += 1
        wr["registered"] += int(rj.get("total_registered", 0) or 0)
        wr["accredited"] += int(rj.get("total_accredited", 0) or 0)
        wr["valid"] += int(rj.get("valid_votes", 0) or 0)
        wr["rejected"] += int(rj.get("invalid_votes", 0) or 0)

        votes_raw = rj.get("votes", "[]")
        if isinstance(votes_raw, str):
            try:
                vote_list = json.loads(votes_raw)
            except:
                vote_list = []
        elif isinstance(votes_raw, list):
            vote_list = votes_raw
        else:
            vote_list = []

        for v in vote_list:
            if isinstance(v, dict):
                party_code = (v.get("party_code", "") or "").upper()
                vote_count = int(v.get("vote", 0) or 0)
                if party_code and vote_count > 0:
                    wr["party_votes"][party_code] = wr["party_votes"].get(party_code, 0) + vote_count

    # Build ward race results
    races = []
    total_party_votes = {}  # Overall party aggregation across all wards
    for eid, wr in ward_races.items():
        party_votes = wr["party_votes"]

        # Sort parties by votes
        sorted_parties = sorted(party_votes.items(), key=lambda x: -x[1])
        candidates = []
        for party, votes in sorted_parties:
            vote_pct = round(votes / wr["valid"] * 100, 1) if wr["valid"] > 0 else 0
            candidates.append({
                "party": party,
                "votes": votes,
                "vote_pct": vote_pct,
            })
            # Aggregate for overall standings
            total_party_votes[party] = total_party_votes.get(party, 0) + votes

        # Margin
        if len(candidates) >= 2 and candidates[0]["votes"] > 0:
            margin = candidates[0]["votes"] - candidates[1]["votes"]
        else:
            margin = 0

        turnout = round(wr["accredited"] / wr["registered"] * 100, 1) if wr["registered"] > 0 else 0
        coverage = round(wr["pus_counted"] / wr["total_pus_in_ward"] * 100, 1) if wr["total_pus_in_ward"] > 0 else 0

        lga_key = wr["lga_name"]
        council_name = LGA_TO_COUNCIL.get(lga_key, lga_key)

        races.append({
            "ward_name": wr["ward_name"],
            "lga_name": wr["lga_name"],
            "area_council": council_name,
            "candidates": candidates[:10],  # Top 10 parties
            "total_pus_in_ward": wr["total_pus_in_ward"],
            "pus_counted": wr["pus_counted"],
            "coverage_pct": coverage,
            "total_valid_votes": wr["valid"],
            "total_registered": wr["registered"],
            "total_accredited": wr["accredited"],
            "total_rejected": wr["rejected"],
            "turnout_pct": turnout,
            "margin": margin,
            "leading_party": candidates[0]["party"] if candidates and candidates[0]["votes"] > 0 else None,
        })

    # Sort by area council then ward name
    races.sort(key=lambda x: (x["area_council"], x["ward_name"]))

    # Overall party standings for councillorship
    party_standings = [
        {"party": p, "votes": v, "wards_leading": 0}
        for p, v in sorted(total_party_votes.items(), key=lambda x: -x[1])
        if v > 0
    ]
    # Count wards each party leads
    for race in races:
        if race["leading_party"]:
            for ps in party_standings:
                if ps["party"] == race["leading_party"]:
                    ps["wards_leading"] += 1
                    break

    # Aggregate by Area Council
    council_summary = {}
    for race in races:
        ac = race["area_council"]
        if ac not in council_summary:
            council_summary[ac] = {
                "area_council": ac,
                "total_wards": 0,
                "wards_with_data": 0,
                "total_valid": 0,
                "total_pus": 0,
                "pus_counted": 0,
                "party_leads": {},
            }
        cs = council_summary[ac]
        cs["total_wards"] += 1
        if race["total_valid_votes"] > 0:
            cs["wards_with_data"] += 1
        cs["total_valid"] += race["total_valid_votes"]
        cs["total_pus"] += race["total_pus_in_ward"]
        cs["pus_counted"] += race["pus_counted"]
        if race["leading_party"]:
            lp = race["leading_party"]
            cs["party_leads"][lp] = cs["party_leads"].get(lp, 0) + 1

    # Also include elections with no PU data yet
    for e in councillor_elections:
        found = False
        for eid_check in ward_races:
            if eid_check == e["id"]:
                found = True
                break
        if not found:
            # This ward has no PU data yet
            ward_name = e["domain_name"] or ""
            # Try to figure out the LGA from existing wards data
            lga_row = conn.execute("""
                SELECT lga_name FROM wards WHERE election_id = ? LIMIT 1
            """, (e["id"],)).fetchone()
            lga_name = lga_row["lga_name"] if lga_row else ""
            council_name = LGA_TO_COUNCIL.get(lga_name, lga_name)

            races.append({
                "ward_name": ward_name,
                "lga_name": lga_name,
                "area_council": council_name,
                "candidates": [],
                "total_pus_in_ward": e["total_pus"],
                "pus_counted": 0,
                "coverage_pct": 0,
                "total_valid_votes": 0,
                "total_registered": 0,
                "total_accredited": 0,
                "total_rejected": 0,
                "turnout_pct": 0,
                "margin": 0,
                "leading_party": None,
            })
            # Also add to council summary
            if council_name:
                if council_name not in council_summary:
                    council_summary[council_name] = {
                        "area_council": council_name,
                        "total_wards": 0, "wards_with_data": 0,
                        "total_valid": 0, "total_pus": 0,
                        "pus_counted": 0, "party_leads": {},
                    }
                council_summary[council_name]["total_wards"] += 1
                council_summary[council_name]["total_pus"] += e["total_pus"]

    races.sort(key=lambda x: (x["area_council"], x["ward_name"]))

    conn.close()

    return jsonify({
        "races": races,
        "party_standings": party_standings,
        "council_summary": sorted(council_summary.values(), key=lambda x: x["area_council"]),
        "total_wards": len(races),
        "wards_with_data": len([r for r in races if r["total_valid_votes"] > 0]),
        "total_councillor_elections": len(councillor_elections),
    })


@app.route("/api/ocr/results")
def api_ocr_results():
    """Legacy OCR results endpoint — now redirects to live-results for better data."""
    return api_live_results()


@app.route("/api/ocr/status")
def api_ocr_status():
    """Quick OCR processing status."""
    conn = get_db()
    stats = conn.execute("""
        SELECT
            COUNT(*) as total_processed,
            SUM(CASE WHEN status='success' THEN 1 ELSE 0 END) as success,
            SUM(CASE WHEN status='low_confidence' THEN 1 ELSE 0 END) as low_confidence,
            SUM(CASE WHEN status='failed' THEN 1 ELSE 0 END) as failed,
            AVG(CASE WHEN status='success' THEN ocr_confidence END) as avg_confidence
        FROM ocr_results
    """).fetchone()
    pending = conn.execute("""
        SELECT COUNT(*) as c FROM polling_units p
        LEFT JOIN ocr_results o ON p.id = o.pu_id
        WHERE p.has_result = 1 AND p.document_url != '' AND o.pu_id IS NULL
    """).fetchone()["c"]
    conn.close()
    return jsonify({
        "total_processed": stats["total_processed"] or 0,
        "success": stats["success"] or 0,
        "low_confidence": stats["low_confidence"] or 0,
        "failed": stats["failed"] or 0,
        "avg_confidence": round(stats["avg_confidence"] or 0, 1),
        "pending": pending,
        "ocr_available": OCR_AVAILABLE,
    })


@app.route("/api/force-ocr", methods=["POST"])
def force_ocr():
    """Manually trigger OCR processing batch."""
    if not OCR_AVAILABLE:
        return jsonify({"error": "OCR not available. Install pytesseract + Pillow."}), 400
    count = process_ocr_batch()
    return jsonify({"message": f"Processed {count} result sheets", "processed": count})


@app.route("/api/force-scrape", methods=["POST"])
def force_scrape():
    if scraper_status["status"] == "scraping":
        return jsonify({"message": "Already scraping"}), 409
    threading.Thread(target=_single_scrape, daemon=True).start()
    return jsonify({"message": "Scrape triggered"})


def _single_scrape():
    try:
        scraper_status["status"] = "scraping"
        scraper_status["error"] = None
        sse_broadcast({"event": "scrape_start", "message": "Scrape triggered"})
        elections = discover_elections()
        scrape_stats(elections)
        scrape_lga_detail(elections)
        scraper_status["last_scrape"] = datetime.now().isoformat()
        scraper_status["status"] = "idle"
        scraper_status["scrape_count"] += 1
        sse_broadcast({"event": "scrape_complete", "scrape_count": scraper_status["scrape_count"]})
    except Exception as e:
        scraper_status["status"] = "error"
        scraper_status["error"] = str(e)
        sse_broadcast({"event": "scrape_error", "error": str(e)})


# ─── SSE (Server-Sent Events) ────────────────────────────────────────────────
def sse_broadcast(data):
    """Push event to all connected SSE clients."""
    with sse_clients_lock:
        dead = []
        for q in sse_clients:
            try:
                q.put_nowait(data)
            except Exception:
                dead.append(q)
        for q in dead:
            sse_clients.remove(q)


@app.route("/api/events")
def api_events():
    """SSE endpoint for real-time scraper updates."""
    def stream():
        q: Queue = Queue()
        with sse_clients_lock:
            sse_clients.append(q)
        try:
            # Send initial status
            yield f"data: {json.dumps({'event': 'connected', **scraper_status})}\n\n"
            while True:
                try:
                    data = q.get(timeout=15)
                    yield f"data: {json.dumps(data)}\n\n"
                except Exception:
                    # Send heartbeat to keep connection alive
                    yield ": heartbeat\n\n"
        except GeneratorExit:
            pass
        finally:
            with sse_clients_lock:
                if q in sse_clients:
                    sse_clients.remove(q)

    return Response(stream(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# ─── Analytics Endpoints ─────────────────────────────────────────────────────
@app.route("/api/analytics/turnout-projection")
def api_turnout_projection():
    """Projected completion based on upload velocity."""
    conn = get_db()
    logs = conn.execute("""
        SELECT timestamp, election_type, results_uploaded, total_pus, percentage
        FROM scrape_log ORDER BY timestamp DESC LIMIT 100
    """).fetchall()
    conn.close()

    if len(logs) < 2:
        return jsonify({"current_rate": 0, "projected_completion": None, "velocity_history": [],
                         "chairman": {"rate": 0, "eta": None}, "councillor": {"rate": 0, "eta": None}})

    logs = [dict(r) for r in reversed(logs)]
    projections = {}

    for etype in ["CHAIRMAN", "COUNCILLOR"]:
        type_logs = [l for l in logs if l["election_type"] == etype]
        if len(type_logs) < 2:
            projections[etype.lower()] = {"rate": 0, "eta": None, "pct": 0}
            continue

        recent = type_logs[-1]
        older = type_logs[-min(6, len(type_logs))]

        try:
            t1 = datetime.fromisoformat(older["timestamp"])
            t2 = datetime.fromisoformat(recent["timestamp"])
            elapsed_min = max((t2 - t1).total_seconds() / 60, 1)
        except Exception:
            elapsed_min = 3

        results_gained = (recent["results_uploaded"] or 0) - (older["results_uploaded"] or 0)
        rate = results_gained / elapsed_min if elapsed_min > 0 else 0
        remaining = (recent["total_pus"] or 0) - (recent["results_uploaded"] or 0)
        eta_min = remaining / rate if rate > 0 else None
        eta_time = None
        if eta_min is not None:
            from datetime import timedelta
            eta_time = (datetime.now() + timedelta(minutes=eta_min)).strftime("%H:%M")

        projections[etype.lower()] = {
            "rate": round(rate, 2),
            "eta": eta_time,
            "remaining": remaining,
            "pct": recent["percentage"] or 0,
        }

    # Velocity history
    velocity = []
    for i in range(1, len(logs)):
        prev, curr = logs[i - 1], logs[i]
        try:
            t1 = datetime.fromisoformat(prev["timestamp"])
            t2 = datetime.fromisoformat(curr["timestamp"])
            elapsed = max((t2 - t1).total_seconds() / 60, 0.1)
            rate = ((curr["results_uploaded"] or 0) - (prev["results_uploaded"] or 0)) / elapsed
            velocity.append({
                "time": curr["timestamp"],
                "rate": round(max(rate, 0), 2),
                "type": curr["election_type"],
            })
        except Exception:
            pass

    return jsonify({
        "chairman": projections.get("chairman", {}),
        "councillor": projections.get("councillor", {}),
        "velocity_history": velocity[-50:],
    })


@app.route("/api/analytics/trends")
def api_analytics_trends():
    """Hourly upload rate trends."""
    conn = get_db()
    logs = conn.execute("""
        SELECT timestamp, election_type, results_uploaded, total_pus, percentage
        FROM scrape_log ORDER BY timestamp ASC
    """).fetchall()
    conn.close()

    hourly = {}
    for log in logs:
        log = dict(log)
        try:
            hour = log["timestamp"][:13] + ":00"
        except Exception:
            continue
        key = (hour, log["election_type"])
        hourly[key] = {
            "hour": hour,
            "type": log["election_type"],
            "results": log["results_uploaded"] or 0,
            "total": log["total_pus"] or 0,
            "pct": log["percentage"] or 0,
        }

    # Build hourly rate deltas
    sorted_keys = sorted(hourly.keys())
    rates = []
    prev_by_type = {}
    for key in sorted_keys:
        entry = hourly[key]
        etype = entry["type"]
        if etype in prev_by_type:
            delta = entry["results"] - prev_by_type[etype]["results"]
            rates.append({
                "hour": entry["hour"],
                "type": etype,
                "uploads": max(delta, 0),
                "pct": entry["pct"],
            })
        prev_by_type[etype] = entry

    # LGA speed ranking from latest data
    conn = get_db()
    lga_speed = conn.execute("""
        SELECT domain_name as lga_name, pct FROM elections
        WHERE election_type='CHAIRMAN' ORDER BY pct DESC
    """).fetchall()
    conn.close()
    lga_speed = [dict(r) for r in lga_speed]

    momentum = "steady"
    if len(rates) >= 4:
        recent_avg = sum(r["uploads"] for r in rates[-3:]) / 3
        older_avg = sum(r["uploads"] for r in rates[-6:-3]) / 3 if len(rates) >= 6 else recent_avg
        if recent_avg > older_avg * 1.2:
            momentum = "accelerating"
        elif recent_avg < older_avg * 0.8:
            momentum = "decelerating"

    return jsonify({
        "hourly_rates": rates[-48:],
        "fastest_lga": lga_speed[0]["lga_name"] if lga_speed else None,
        "slowest_lga": lga_speed[-1]["lga_name"] if lga_speed else None,
        "lga_ranking": lga_speed,
        "momentum": momentum,
    })


@app.route("/api/analytics/heatmap")
def api_analytics_heatmap():
    """Ward-level result percentages for heat map."""
    conn = get_db()
    wards = conn.execute("""
        SELECT w.ward_name, w.lga_name, w.total_pus, w.results_uploaded,
               CASE WHEN w.total_pus > 0 THEN ROUND(w.results_uploaded * 100.0 / w.total_pus, 1) ELSE 0 END as pct
        FROM wards w
        ORDER BY w.lga_name, w.ward_name
    """).fetchall()

    # Also include LGA-level from elections table
    lgas = conn.execute("""
        SELECT domain_name as lga_name, total_pus, total_results as results_uploaded,
               pct FROM elections WHERE election_type='CHAIRMAN' ORDER BY domain_name
    """).fetchall()
    conn.close()

    return jsonify({
        "wards": [dict(r) for r in wards],
        "lgas": [dict(r) for r in lgas],
    })


# ─── Export Endpoints ─────────────────────────────────────────────────────────
def _make_csv(rows, fieldnames):
    """Generate CSV string from list of dicts."""
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return output.getvalue()


def _make_xlsx(rows, fieldnames, sheet_name="Data"):
    """Generate XLSX bytes from list of dicts."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(fieldnames)
    for row in rows:
        ws.append([row.get(f, "") for f in fieldnames])
    # Auto-width columns
    for col_idx, field in enumerate(fieldnames, 1):
        max_len = max(len(str(field)), *(len(str(row.get(field, ""))) for row in rows)) if rows else len(str(field))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 40)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.route("/api/export/elections")
def api_export_elections():
    fmt = request.args.get("format", "csv")
    if fmt not in ("csv", "xlsx"):
        return jsonify({"error": "Invalid format. Use 'csv' or 'xlsx'"}), 400
    try:
        conn = get_db()
        rows = [dict(r) for r in conn.execute("""
            SELECT domain_name as area_council, election_type, total_pus, total_results,
                   pct as percentage, election_date, last_updated
            FROM elections ORDER BY election_type, domain_name
        """).fetchall()]
        conn.close()

        fields = ["area_council", "election_type", "total_pus", "total_results", "percentage", "election_date", "last_updated"]

        if fmt == "xlsx":
            output = _make_xlsx(rows, fields, "Elections")
            return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             as_attachment=True, download_name="elections_export.xlsx")
        else:
            csv_str = _make_csv(rows, fields)
            return Response(csv_str, mimetype="text/csv",
                            headers={"Content-Disposition": "attachment; filename=elections_export.csv"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/export/candidates")
def api_export_candidates():
    fmt = request.args.get("format", "csv")
    if fmt not in ("csv", "xlsx"):
        return jsonify({"error": "Invalid format. Use 'csv' or 'xlsx'"}), 400
    try:
        conn = get_db()
        rows = [dict(r) for r in conn.execute(
            "SELECT area_council, candidate_name, party_full, party_abbrev, status, gender, position_type FROM candidates ORDER BY area_council, party_abbrev"
        ).fetchall()]
        conn.close()

        fields = ["area_council", "candidate_name", "party_full", "party_abbrev", "status", "gender", "position_type"]

        if fmt == "xlsx":
            output = _make_xlsx(rows, fields, "Candidates")
            return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             as_attachment=True, download_name="candidates_export.xlsx")
        else:
            csv_str = _make_csv(rows, fields)
            return Response(csv_str, mimetype="text/csv",
                            headers={"Content-Disposition": "attachment; filename=candidates_export.csv"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/export/analytics")
def api_export_analytics():
    fmt = request.args.get("format", "csv")
    if fmt not in ("csv", "xlsx"):
        return jsonify({"error": "Invalid format. Use 'csv' or 'xlsx'"}), 400
    try:
        conn = get_db()
        rows = [dict(r) for r in conn.execute("""
            SELECT timestamp, election_type, total_pus, results_uploaded, percentage
            FROM scrape_log ORDER BY timestamp DESC LIMIT 500
        """).fetchall()]
        conn.close()

        fields = ["timestamp", "election_type", "total_pus", "results_uploaded", "percentage"]

        if fmt == "xlsx":
            output = _make_xlsx(rows, fields, "Analytics")
            return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             as_attachment=True, download_name="analytics_export.xlsx")
        else:
            csv_str = _make_csv(rows, fields)
            return Response(csv_str, mimetype="text/csv",
                            headers={"Content-Disposition": "attachment; filename=analytics_export.csv"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Messaging Stats ─────────────────────────────────────────────────────────
@app.route("/api/messaging/stats")
def api_messaging_stats():
    """Read WhatsApp send progress from JSON files."""
    base = os.path.dirname(os.path.abspath(__file__))
    stats = {"sent": 0, "failed": 0, "total_agents": 0}

    for fname in ["sms_send_progress_v2.json", "sms_send_progress.json"]:
        fpath = os.path.join(base, fname)
        if os.path.exists(fpath):
            try:
                with open(fpath, "r") as f:
                    data = json.load(f)
                stats["sent"] = len(data.get("sent", []))
                stats["failed"] = len(data.get("failed", []))
                stats["sent_numbers"] = data.get("sent", [])
                stats["failed_numbers"] = data.get("failed", [])
                break
            except Exception:
                pass

    # Count CSV agents if available
    csv_path = os.path.join(base, "polling_agents_1771679503772.csv")
    if os.path.exists(csv_path):
        try:
            with open(csv_path, "r") as f:
                reader = csv.DictReader(f)
                agents = list(reader)
                stats["total_agents"] = len(agents)
        except Exception:
            pass

    stats["pending"] = max(0, stats["total_agents"] - stats["sent"] - stats["failed"])
    return jsonify(stats)


# ─── Module-level init (runs under gunicorn AND direct python) ────────────────
_initialized = False

def _ensure_init():
    global _initialized
    if _initialized:
        return
    _initialized = True
    print("=" * 60)
    print("  FCT 2026 Area Council Elections - Live Dashboard")
    print("  Date: Saturday, 21 February 2026")
    print("=" * 60)
    init_db()
    load_excel_data()
    scraper_thread = threading.Thread(target=scrape_loop, daemon=True)
    scraper_thread.start()
    print("[OK] Background scraper started (every 2 minutes)")

_ensure_init()

# ─── Main (local dev) ─────────────────────────────────────────────────────────
if __name__ == "__main__":
    port = int(os.getenv("PORT", 5050))
    print(f"[OK] Dashboard: http://localhost:{port}")
    print("=" * 60)
    app.run(host="0.0.0.0", port=port, debug=False, threaded=True)
